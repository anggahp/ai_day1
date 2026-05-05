import json
import csv
import os
from typing import Type
from crewai.tools import BaseTool
from pydantic import BaseModel, Field


class FileReaderInput(BaseModel):
    """Input schema for FileReaderTool."""
    file_path: str = Field(..., description="Absolute or relative path to the file (TXT, JSON, CSV, or Excel)")


class FileReaderTool(BaseTool):
    name: str = "file_reader"
    description: str = (
        "Reads the content of a file. Supports TXT, JSON, CSV, and Excel (.xlsx/.xls) formats. "
        "For JSON, it pretty-prints the structure. "
        "For CSV, it converts rows to a readable table format. "
        "For Excel, it reads all sheets and presents data as a table. "
        "Returns the file content as a string."
    )
    args_schema: Type[BaseModel] = FileReaderInput

    def _run(self, file_path: str) -> str:
        if not os.path.exists(file_path):
            return f"Error: File not found at path '{file_path}'"

        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == ".json":
                return self._read_json(file_path)
            elif ext == ".csv":
                return self._read_csv(file_path)
            elif ext in (".xlsx", ".xls"):
                return self._read_excel(file_path)
            else:
                return self._read_txt(file_path)
        except Exception as e:
            return f"Error reading file: {str(e)}"

    def _read_txt(self, file_path: str) -> str:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return f"[TXT File Content]\n{content}"

    def _read_json(self, file_path: str) -> str:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            total_rows = len(data)
            # Show summary + first 20 rows to avoid token overflow
            preview = data[:20]
            formatted = json.dumps(preview, indent=2, ensure_ascii=False)
            summary = f"[JSON File — {total_rows} records total, showing first {len(preview)}]\n"
            if total_rows > 20:
                summary += f"(Remaining {total_rows - 20} records not shown)\n"
            return summary + formatted
        else:
            formatted = json.dumps(data, indent=2, ensure_ascii=False)
            return f"[JSON File Content]\n{formatted}"

    def _read_csv(self, file_path: str) -> str:
        rows = []
        with open(file_path, "r", encoding="utf-8") as f:
            # Detect delimiter (tab or comma)
            sample = f.read(1024)
            f.seek(0)
            delimiter = "\t" if "\t" in sample else ","
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                rows.append(dict(row))

        total_rows = len(rows)
        preview = rows[:20]
        headers = list(preview[0].keys()) if preview else []

        # Format as readable table
        lines = ["\t".join(headers)]
        for row in preview:
            lines.append("\t".join(str(row.get(h, "")) for h in headers))

        summary = f"[CSV File — {total_rows} rows, {len(headers)} columns, showing first {len(preview)}]\n"
        if total_rows > 20:
            summary += f"(Remaining {total_rows - 20} rows not shown)\n"
        summary += f"Columns: {', '.join(headers)}\n\n"
        return summary + "\n".join(lines)

    def _read_excel(self, file_path: str) -> str:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        output_parts = []
        sheet_names = wb.sheetnames
        output_parts.append(f"[Excel File — {len(sheet_names)} sheet(s): {', '.join(sheet_names)}]")

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            if not all_rows:
                output_parts.append(f"\n--- Sheet: '{sheet_name}' (empty) ---")
                continue

            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            data_rows = all_rows[1:]
            total_rows = len(data_rows)
            preview = data_rows[:20]

            output_parts.append(f"\n--- Sheet: '{sheet_name}' — {total_rows} rows, {len(headers)} columns ---")
            if total_rows > 20:
                output_parts.append(f"(Showing first 20 of {total_rows} rows)")
            output_parts.append(f"Columns: {', '.join(headers)}\n")

            # Format as readable table
            lines = ["\t".join(headers)]
            for row in preview:
                lines.append("\t".join(str(v) if v is not None else "" for v in row))
            output_parts.append("\n".join(lines))

        return "\n".join(output_parts)
